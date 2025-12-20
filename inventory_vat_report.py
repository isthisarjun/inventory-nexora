import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "VAT_Report_Data"

# Column headers
headers = [
    "transaction_id",
    "transaction_date",
    "transaction_type",
    "document_number",
    "party_name",
    "description",
    "net_amount",
    "vat_amount",
    "gross_amount",
    "is_vat_applicable",
    "is_recoverable",
    "vat_period",
    "reporting_year",
    "adjustment_type",
    "related_document",
    "adjustment_reason",
]
ws.append(headers)

# Data validation rules
transaction_type_dv = DataValidation(type="list", formula1='"SALE,PURCHASE,ADJUSTMENT"', allow_blank=True)
adjustment_type_dv = DataValidation(type="list", formula1='"CREDIT,DEBIT,NONE"', allow_blank=True)
true_false_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)

# Add validations to columns
ws.add_data_validation(transaction_type_dv)
ws.add_data_validation(adjustment_type_dv)
ws.add_data_validation(true_false_dv)

# Apply validations to columns (2nd row onwards)
for row in range(2, 1002):
    transaction_type_dv.add(ws[f"C{row}"])
    adjustment_type_dv.add(ws[f"N{row}"])
    true_false_dv.add(ws[f"J{row}"])
    true_false_dv.add(ws[f"K{row}"])

# Lock vat_amount and gross_amount columns
for col in [8, 9]:  # H and I
    for row in range(2, 1002):
        ws.cell(row=row, column=col).protection = openpyxl.styles.Protection(locked=True)

# Set formulas for vat_amount and gross_amount
for row in range(2, 1002):
    # vat_amount: IF(is_vat_applicable = TRUE, net_amount * 0.10, 0)
    ws[f"H{row}"].value = f'=IF(J{row}="TRUE", G{row}*0.10, 0)'
    # gross_amount: net_amount + vat_amount
    ws[f"I{row}"].value = f'=G{row}+H{row}'

# Format date columns (transaction_date, vat_period)
date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
for col in [2, 12]:  # B and L
    for row in range(2, 1002):
        ws.cell(row=row, column=col).style = date_style

# Format amount columns to 3 decimal places (net_amount, vat_amount, gross_amount)
amount_style = NamedStyle(name="amount_style", number_format="#.000")
for col in [7, 8, 9]:  # G, H, I
    for row in range(2, 1002):
        ws.cell(row=row, column=col).style = amount_style

# Set column widths for readability
for i, header in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(i)].width = 18

# Protect sheet to enforce locked cells
ws.protection.sheet = True
ws.protection.enable()

# Save file
wb.save("inventory_vat_report.xlsx")
